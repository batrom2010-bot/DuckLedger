import asyncio
import logging
import os
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Dict, List, Tuple

from aiogram import Bot, Dispatcher, F, Router
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    CallbackQuery,
    FSInputFile,
    Message,
)
from openpyxl import Workbook
from pathlib import Path

# ===============================
# НАСТРОЙКИ
# ===============================

DB_FILE = "budget.db"

# Жёстко считаем локальное время = UTC+3
LOCAL_UTC_OFFSET = 3  # часы


def get_local_now() -> datetime:
    """Текущее локальное время (UTC+3), без таймзоны."""
    return datetime.utcnow() + timedelta(hours=LOCAL_UTC_OFFSET)


BOT_TOKEN = os.getenv("BOT_TOKEN") or os.getenv("TELEGRAM_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("Укажи токен бота в BOT_TOKEN или переменной TELEGRAM_TOKEN")

# ===============================
# ЛОГИ
# ===============================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

# ===============================
# FSM
# ===============================


class InsertStates(StatesGroup):
    waiting_for_expenses = State()


class LimitStates(StatesGroup):
    waiting_for_limits = State()


# ===============================
# РАБОТА С БАЗОЙ
# ===============================


def init_db() -> None:
    """Создаём таблицы, если их ещё нет."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    # Таблица расходов
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            amount REAL NOT NULL,
            timestamp TEXT NOT NULL
        )
        """
    )

    # Таблица лимитов по категориям
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS limits (
            category TEXT PRIMARY KEY,
            limit_amount REAL NOT NULL
        )
        """
    )

    conn.commit()
    conn.close()
    logger.info("Инициализация базы данных...")


from datetime import datetime, timezone, timedelta

LOCAL_TZ = timezone(timedelta(hours=3))  # Москва / твой часовой пояс

def add_expense(category: str, amount: float):
    """Добавить расход в БД."""
    # Дата в локальном поясе, формат как в Excel: 14/11/2025
    ts = datetime.now(LOCAL_TZ).strftime("%d/%m/%Y")

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO expenses (category, amount, timestamp)
        VALUES (?, ?, ?)
        """,
        (category, amount, ts),
    )
    conn.commit()
    conn.close()
    logger.info(f"Добавлен расход: {category} - {amount}, дата: {ts}")



def set_limits(pairs: List[Tuple[str, float]]) -> None:
    """Установить лимиты по категориям (upsert)."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    for category, limit_amount in pairs:
        cursor.execute(
            """
            INSERT INTO limits (category, limit_amount)
            VALUES (?, ?)
            ON CONFLICT(category) DO UPDATE SET limit_amount=excluded.limit_amount
            """,
            (category.strip(), limit_amount),
        )
    conn.commit()
    conn.close()


@dataclass
class MonthStats:
    total: float
    by_category: Dict[str, float]
    limits: Dict[str, float]


def _load_all_expenses() -> List[Tuple[str, str, float]]:
    """Читаем все расходы: (timestamp_str, category, amount)."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT timestamp, category, amount FROM expenses")
    rows = cursor.fetchall()
    conn.close()
    return [(r[0], r[1], float(r[2])) for r in rows]


def _load_limits() -> Dict[str, float]:
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT category, limit_amount FROM limits")
    rows = cursor.fetchall()
    conn.close()
    return {r[0]: float(r[1]) for r in rows}


def get_month_stats() -> MonthStats:
    """Краткая статистика за текущий месяц по локальному времени."""
    now = get_local_now()
    year = now.year
    month = now.month

    rows = _load_all_expenses()
    by_cat: Dict[str, float] = {}
    total = 0.0

    for ts_str, category, amount in rows:
        try:
            dt = datetime.fromisoformat(ts_str)
        except ValueError:
            continue
        if dt.year == year and dt.month == month:
            by_cat[category] = by_cat.get(category, 0.0) + amount
            total += amount

    limits = _load_limits()
    return MonthStats(total=total, by_category=by_cat, limits=limits)


def get_full_stats() -> MonthStats:
    """Статистика за весь период (используем ту же структуру)."""
    rows = _load_all_expenses()
    by_cat: Dict[str, float] = {}
    total = 0.0

    for _, category, amount in rows:
        by_cat[category] = by_cat.get(category, 0.0) + amount
        total += amount

    limits = _load_limits()
    return MonthStats(total=total, by_category=by_cat, limits=limits)


def export_to_excel(file_path: str):
    """
    Экспорт расходов в Excel.

    Формат:
    Дата | Категория1 | Категория2 | ...
    14/11/2025 | 1000 | 500 | ...
    """
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT timestamp, category, amount FROM expenses ORDER BY timestamp"
    )
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        return False

    # Собираем все даты и категории
    dates = sorted({row[0] for row in rows})
    categories = sorted({row[1] for row in rows})

    # Агрегируем суммы: дата -> категория -> сумма
    data = {d: {c: 0 for c in categories} for d in dates}
    for ts, cat, amount in rows:
        data[ts][cat] += amount

    wb = Workbook()
    ws = wb.active
    ws.title = "Расходы"

    # Заголовки
    ws.cell(row=1, column=1, value="Дата")
    for col, cat in enumerate(categories, start=2):
        ws.cell(row=1, column=col, value=cat)

    # Строки по датам
    for row_idx, d in enumerate(dates, start=2):
        ws.cell(row=row_idx, column=1, value=d)
        for col_idx, cat in enumerate(categories, start=2):
            val = data[d][cat]
            if val:  # пустые не пишем, чтобы не засорять нулями
                ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(file_path)
    return True


    # Собираем все даты и категории
    data: Dict[datetime.date, Dict[str, float]] = {}
    categories_set = set()

    for ts_str, category, amount in rows:
        try:
            dt = datetime.fromisoformat(ts_str)
        except ValueError:
            continue
        date_key = dt.date()  # уже локальная дата (мы писали локальное время)
        if date_key not in data:
            data[date_key] = {}
        data[date_key][category] = data[date_key].get(category, 0.0) + amount
        categories_set.add(category)

    categories = sorted(categories_set)
    dates_sorted = sorted(data.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Расходы"

    # Шапка
    header = ["Дата"] + categories
    ws.append(header)

    # Строки
    for d in dates_sorted:
        row = [d.strftime("%d/%m/%Y")]
        for cat in categories:
            value = data[d].get(cat, 0.0)
            row.append(value if value != 0.0 else "")
        ws.append(row)

    wb.save(file_path)


def clear_expenses() -> None:
    """Полностью очищаем таблицу расходов."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM expenses")
    conn.commit()
    conn.close()


# ===============================
# УТИЛИТЫ ПАРСИНГА
# ===============================


def parse_lines_to_pairs(text: str) -> List[Tuple[str, float]]:
    """
    Разбирает список строк формата:

    Категория-Сумма
    Категория2-100

    Возвращает список (category, amount).
    """
    pairs: List[Tuple[str, float]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if "-" not in line:
            raise ValueError(
                f"Не удалось найти разделитель '-' в строке: «{line}»\n"
                f"Пример: Еда-500"
            )
        category, amount_str = line.split("-", 1)
        category = category.strip()
        amount_str = amount_str.strip().replace(",", ".")
        if not category:
            raise ValueError(f"Пустая категория в строке: «{line}»")
        try:
            amount = float(amount_str)
        except ValueError:
            raise ValueError(f"Не получилось прочитать сумму в строке: «{line}»")
        pairs.append((category, amount))
    return pairs


# ===============================
# TELEGRAM
# ===============================

router = Router()


@router.message(Command("start"))
async def cmd_start(message: Message) -> None:
    text = (
        "Привет, ! Я бот для учёта расходов.\n\n"
        "Просто отправь мне строку в формате:\n"
        "<b>Категория-Сумма</b>\n"
        "или используй команду <b>/insert</b> для ввода списка.\n\n"
        "Команда <b>/limit</b> — для установки лимитов по категориям.\n"
        "Команда <b>/stats</b> — краткая статистика по месяцу.\n"
        "Команда <b>/analitick</b> — расширенная аналитика.\n"
        "Команда <b>/make</b> — сформировать Excel-отчёт и очистить базу.\n"
        "Команда <b>/export</b> — выгрузить Excel-таблицу (без очистки).\n"
    )
    await message.answer(text)


@router.message(Command("help"))
async def cmd_help(message: Message) -> None:
    text = (
        "<b>Формат ввода расходов</b>\n"
        "Одна строка:\n"
        "  <code>Еда-500</code>\n\n"
        "Список через /insert:\n"
        "  <code>Еда-500\nТакси-300\nКофе-200</code>\n\n"
        "<b>/limit</b> — установить лимиты по категориям.\n"
        "Формат такой же, можно несколько строк.\n\n"
        "<b>/stats</b> — расходы по текущему месяцу с учётом лимитов.\n"
        "<b>/analitick</b> — суммарная аналитика за весь период.\n"
        "<b>/export</b> — Excel с агрегированными данными (дата x категория).\n"
        "<b>/make</b> — такой же Excel + очистка базы. Сохрани файл у себя, "
        "после команды /make данные будут удалены.\n"
    )
    await message.answer(text)


# ---------- /insert ----------


@router.message(Command("insert"))
async def cmd_insert(message: Message, state: FSMContext) -> None:
    text = (
        "Отправь список расходов в формате:\n"
        "<b>Категория-Сумма</b>\n"
        "Можно сразу несколько строк:\n"
        "<code>Еда-500\nТакси-300\nКофе-200</code>"
    )
    await state.set_state(InsertStates.waiting_for_expenses)
    await message.answer(text)


@router.message(InsertStates.waiting_for_expenses)
async def process_insert_list(message: Message, state: FSMContext) -> None:
    try:
        pairs = parse_lines_to_pairs(message.text or "")
    except ValueError as e:
        await message.answer(f"⚠️ Ошибка: {e}")
        return

    for category, amount in pairs:
        add_expense(category, amount)

    await state.clear()
    total = sum(a for _, a in pairs)
    await message.answer(
        f"Записал {len(pairs)} расходов на сумму {total:.0f}. "
        f"Можешь отправлять новые строки в формате <b>Категория-Сумма</b>."
    )


# ---------- /limit ----------


@router.message(Command("limit"))
async def cmd_limit(message: Message, state: FSMContext) -> None:
    text = (
        "Отправь список лимитов в формате:\n"
        "<b>Категория-Сумма</b>\n"
        "Можно сразу несколько строк:\n"
        "<code>Еда-20000\nТакси-5000</code>"
    )
    await state.set_state(LimitStates.waiting_for_limits)
    await message.answer(text)


@router.message(LimitStates.waiting_for_limits)
async def process_limit_list(message: Message, state: FSMContext) -> None:
    try:
        pairs = parse_lines_to_pairs(message.text or "")
    except ValueError as e:
        await message.answer(f"⚠️ Ошибка: {e}")
        return

    set_limits(pairs)
    await state.clear()
    await message.answer("Лимиты обновлены.")


# ---------- /stats ----------


@router.message(Command("stats"))
async def cmd_stats(message: Message) -> None:
    stats = get_month_stats()
    if not stats.by_category:
        await message.answer("За этот месяц расходов ещё нет.")
        return

    lines = ["<b>Расходы за текущий месяц:</b>"]
    for cat, amount in sorted(stats.by_category.items()):
        line = f"{cat}: {amount:.0f}"
        if cat in stats.limits:
            limit = stats.limits[cat]
            diff = limit - amount
            if diff < 0:
                line += f" (перерасход {abs(diff):.0f} от лимита {limit:.0f})"
            else:
                line += f" (остаток {diff:.0f} из лимита {limit:.0f})"
        lines.append(line)

    lines.append(f"\n<b>Всего:</b> {stats.total:.0f}")
    await message.answer("\n".join(lines))


# ---------- /analitick ----------


@router.message(Command("analitick"))
async def cmd_analitick(message: Message) -> None:
    stats = get_full_stats()
    if not stats.by_category:
        await message.answer("Расходов пока нет.")
        return

    lines = ["<b>Общая аналитика по всем расходам:</b>"]
    for cat, amount in sorted(stats.by_category.items(), key=lambda x: -x[1]):
        lines.append(f"{cat}: {amount:.0f}")

    lines.append(f"\n<b>Всего за всё время:</b> {stats.total:.0f}")
    await message.answer("\n".join(lines))


# ---------- /export & /make ----------


async def _send_excel(message: Message, clear_after: bool) -> None:
    rows = _load_all_expenses()
    if not rows:
        await message.answer("Пока нет данных для отчёта.")
        return

    file_path = "budget_export.xlsx"
    export_to_excel(file_path)

    doc = FSInputFile(file_path)
    caption = "Excel-отчёт по расходам."
    if clear_after:
        caption += "\nПосле этой команды база будет очищена."

    await message.answer_document(doc, caption=caption)

    if clear_after:
        clear_expenses()
        await message.answer(
            "Я обнулил базу расходов. Сохрани файл Excel у себя, "
            "новые записи будут накапливаться с нуля."
        )


@router.message(Command("export"))
async def cmd_export(message: Message):
    tmp_dir = Path("export")
    tmp_dir.mkdir(exist_ok=True)
    file_path = tmp_dir / "expenses.xlsx"

    if not export_to_excel(str(file_path)):
        await message.answer("Нет данных для экспорта.")
        return

    await message.answer_document(
        FSInputFile(str(file_path)),
        caption="Вот твоя таблица расходов 📊",
    )



@router.message(Command("make"))
async def cmd_make(message: Message):
    """
    1) Выгружает Excel в агрегированном формате (как /export).
    2) Очищает таблицу expenses.
    3) Предупреждает, что данные сброшены.
    """
    tmp_dir = Path("export")
    tmp_dir.mkdir(exist_ok=True)
    file_path = tmp_dir / "expenses_make.xlsx"

    if not export_to_excel(str(file_path)):
        await message.answer("Нет данных для формирования отчёта.")
        return

    await message.answer_document(
        FSInputFile(str(file_path)),
        caption=(
            "Отчёт сформирован ✅\n\n"
            "Файл содержит ВСЕ текущие данные. "
            "После этого база будет очищена."
        ),
    )

    # Чистим базу
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM expenses")
    conn.commit()
    conn.close()

    await message.answer(
        "Данные в боте очищены. Сохрани файл, если он тебе нужен 📁"
    )


# ---------- одиночная строка "Категория-Сумма" ----------


@router.message(F.text)
async def process_single_line(message: Message) -> None:
    text = (message.text or "").strip()
    if not text:
        return
    try:
        pairs = parse_lines_to_pairs(text)
    except ValueError:
        # Не наш формат — просто игнорируем/не ломаемся.
        return

    # берём только первую пару, так как это обычное сообщение
    category, amount = pairs[0]
    add_expense(category, amount)
    await message.answer(f"Записал: {category} — {amount:.0f}")


# ===============================
# MAIN
# ===============================


async def main() -> None:
    logger.info("Запускаем DuckLedger...")
    init_db()

    bot = Bot(token=BOT_TOKEN, parse_mode=ParseMode.HTML)
    dp = Dispatcher()
    dp.include_router(router)

    # long polling
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())






